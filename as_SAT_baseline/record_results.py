import pandas as pd
import numpy as np
import openpyxl
import json


def merge(mod_label_json, xlsx2load, xlsx2save):
    dataset2mod = {}
    mod_lab2dice = {}
    
    # Load the first sheet of the Excel file
    excel_file_path = xlsx2load 
    df = pd.read_excel(excel_file_path, sheet_name=0, engine='openpyxl')
    has_nsd = True if len(df.columns) > 2 else False
    
    # 将Dataset Merged 写入新的工作表
    workbook = openpyxl.load_workbook(xlsx2load)
    new_sheet = workbook.create_sheet(title='Dataset Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Dataset')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    row = 2
    for i in range(0, len(df)):
        if ',' not in df.iloc[i, 0]:
            new_sheet.cell(row=row, column=1, value=df.iloc[i, 0])
            new_sheet.cell(row=row, column=2, value=df.iloc[i, 1])
            if has_nsd:
                new_sheet.cell(row=row, column=3, value=df.iloc[i, 2])
            row += 1
    
    # with pd.ExcelWriter(xlsx2save, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    #     filtered_df.to_excel(writer, sheet_name='Dataset Merge', index=False)
    
    # 选取前两列
    dataset_label_ls = df.iloc[:, 0]
    dice_ls = df.iloc[:, 1]
    nsd_ls = df.iloc[:, 2] if has_nsd else [0] * len(df)
    
    for dataset_modality_label, dice, nsd in zip(dataset_label_ls, dice_ls, nsd_ls):  # MSD_Pancreas(ct), pancreas   0.89
        if ', ' not in dataset_modality_label:
            continue
        dataset_modality, label = dataset_modality_label.split(', ')
        label = label.lower()   # pancreas
        # label = merge_label(label)
        modality = dataset_modality.split('(')[-1].split(')')[0]   # ct
            
        # unique id : modality_label
        mod_lab = f'{modality}_{label}'
        
        # accumulate : dice and where the dice comes from (dataset, label, modality)
        if mod_lab not in mod_lab2dice:
            mod_lab2dice[mod_lab] = {'dice':[], 'nsd':[], 'merge':[]}
        mod_lab2dice[mod_lab]['dice'].append(dice)
        mod_lab2dice[mod_lab]['nsd'].append(nsd)
        mod_lab2dice[mod_lab]['merge'].append(dataset_modality_label)
        
    # retrieval regions
    with open(mod_label_json, 'r') as f:
        dict = json.load(f)
    region2label = dict['region_based']
    for region, label_ls in region2label.items():
        region2label[region] = [mod_lab.split('_')[-1] for mod_lab in label_ls] # 去除modality
    region2label['abnormal'] = [mod_lab.split('_')[-1] for mod_lab in dict['abnormal']]
        
    region_dice_ls = {k:[] for k in region2label.keys()} # {'brain':[0.9, ...], ...}
    region_nsd_ls = {k:[] for k in region2label.keys()} # {'brain':[0.9, ...], ...}
    region_merge_ls = {k:[] for k in region2label.keys()} # {'brain':['frontal lobe', ...], ...}
    
    mod_lab_ls = []
    dice_ls = []
    nsd_ls = []
    merge_ls = []  
    region_ls = []
    for mod_lab, dict in mod_lab2dice.items():
        label = mod_lab.split('_')[-1]
        mod_lab_ls.append(mod_lab)
        dice_ls.append(sum(dict['dice'])/len(dict['dice']))
        nsd_ls.append(sum(dict['nsd'])/len(dict['nsd']))
        merge_ls.append(' / '.join(dict['merge']))
        
        # find region
        if label in region2label['abnormal']:
            region_dice_ls['abnormal'].append(dice_ls[-1])
            region_nsd_ls['abnormal'].append(nsd_ls[-1])
            region_merge_ls['abnormal'].append(mod_lab)
            region_ls.append('abnormal')
        else:
            found = False
            for region, labels_in_region in region2label.items():
                if label in labels_in_region:
                    region_dice_ls[region].append(dice_ls[-1])
                    region_nsd_ls[region].append(nsd_ls[-1])
                    region_merge_ls[region].append(mod_lab)
                    region_ls.append(region)
                    found = True
                    break
            if not found:
                print(label)
                region_ls.append('unknown')
            
    df = pd.DataFrame({
        'Modality_Label': mod_lab_ls,
        'Dice': dice_ls,
        'NSD': nsd_ls,
        'Merge': merge_ls,
        'Region': region_ls
    })

    #book = openpyxl.load_workbook(xlsx2save)
    #writer = pd.ExcelWriter(xlsx2save, engine='openpyxl')
    #writer.book = book
    
    # with pd.ExcelWriter(xlsx2save, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    #     df.to_excel(writer, sheet_name='Label Merge', index=False)

    # 将Label Merged DataFrame写入新的工作表
    new_sheet = workbook.create_sheet(title='Label Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Modality_Label')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    new_sheet.cell(row=1, column=4, value='Merge')
    new_sheet.cell(row=1, column=5, value='Region')
    row = 2
    for mod_lab, dice, nsd, merge, region in zip(mod_lab_ls, dice_ls, nsd_ls, merge_ls, region_ls):
        new_sheet.cell(row=row, column=1, value=mod_lab)
        new_sheet.cell(row=row, column=2, value=dice)
        new_sheet.cell(row=row, column=3, value=nsd)
        new_sheet.cell(row=row, column=4, value=merge)
        new_sheet.cell(row=row, column=5, value=region)
        row += 1
    new_sheet.cell(row=row, column=2, value=sum(dice_ls)/len(dice_ls))  # avg over all labels
    new_sheet.cell(row=row, column=3, value=sum(nsd_ls)/len(nsd_ls))
    
    # 将Region Merged 写入新的工作表
    new_sheet = workbook.create_sheet(title='Region Merge', index=1)
    new_sheet.cell(row=1, column=1, value='Region')
    new_sheet.cell(row=1, column=2, value='Dice')
    new_sheet.cell(row=1, column=3, value='NSD')
    new_sheet.cell(row=1, column=4, value='Merge')
    row = 2
    for key in region_dice_ls.keys():
        if len(region_dice_ls[key]) == 0:
            dice = nsd = 0
            merge = None
        else:
            dice = sum(region_dice_ls[key])/len(region_dice_ls[key])
            nsd = sum(region_nsd_ls[key])/len(region_nsd_ls[key])
            merge = ','.join(region_merge_ls[key])
        class_name = f'{key}({len(region_dice_ls[key])})'
        new_sheet.cell(row=row, column=1, value=class_name)
        new_sheet.cell(row=row, column=2, value=dice)
        new_sheet.cell(row=row, column=3, value=nsd)
        new_sheet.cell(row=row, column=4, value=merge)
        row += 1

    workbook.save(xlsx2save)
    
    # 返回所有 label 的 avg
    avg_dice_over_merged_labels = sum(dice_ls) / len(dice_ls)
    avg_nsd_over_merged_labels = sum(nsd_ls) / len(nsd_ls)
    
    return avg_dice_over_merged_labels, avg_nsd_over_merged_labels


def record(results_of_samples, csv_path):
    # datasets --> labels --> metrics
    datasets_labels_metrics = {}   # {'COVID19':{'covid19_infection':{'dice':[0.8, 0.9, ...], ...} ...}, ...}

    # datasets --> samples --> labels --> metrics
    samples_labels_metrics = {}   # {'COVID19':{'0.npy':{'covid19_infection':{'dice':0.8, ...} ...}, ...} 记录每个dataset里的sample（行）
    
    # datsets --> labels
    datasets_labels_sets = {}    # {'COVID19':set('covid19_infection', ...), ...}  记录每个dataset里的label种类（列）
            
    # collate results
    for dataset_name, modality, sample_id, scores, labels in results_of_samples:    #  [[dataset_name, modality, sample_id, [{'dice':x, ...}, ...], ['liver', ...]], ...]
        dataset_name = f'{dataset_name}({modality})'
        
        if dataset_name not in datasets_labels_metrics:
            datasets_labels_metrics[dataset_name] = {}  # {'COVID19(CT)':{}}
        if dataset_name not in datasets_labels_sets:
            datasets_labels_sets[dataset_name] = set()  # {'COVID19(CT)':set()}
        if dataset_name not in samples_labels_metrics:
            samples_labels_metrics[dataset_name] = {}
        samples_labels_metrics[dataset_name][sample_id] = {}   # {'COVID19(CT)':{'0':{}}}
        
        for metric_dict, label in zip(scores, labels):
            # accumulate metrics （for per dataset per class
            # {'COVID19(CT)':{'covid19_infection':{'dice':[0.8, 0.9, ...], 'nsd':[0.8, 0.9, ...], ...} ...}, ...}
            if label not in datasets_labels_metrics[dataset_name]:
                datasets_labels_metrics[dataset_name][label] = {k:[v] for k,v in metric_dict.items()}
            else:
                for k,v in metric_dict.items():
                    datasets_labels_metrics[dataset_name][label][k].append(v)
            
            # statistic labels
            # {'COVID19(CT)':set('covid19_infection', ...)}
            if label not in datasets_labels_sets[dataset_name]:
                datasets_labels_sets[dataset_name].add(label)
            
            # record metrics （for per dataset per sample per class
            # {'COVID19':{'0.npy':{'covid19_infection':{'dice':0.8, 'nsd':0.9, ...} ...}, ...}
            samples_labels_metrics[dataset_name][sample_id][label] = {k:v for k,v in metric_dict.items()}
                
    # average and log (列为metrics，例如dice，nsd...)
    # create a df like:
    # {
    #   'TotalSegmentator': [0.xx, 0.xx, ...]    # 在T之前，这是一列
    #   'TotalSegmentator, Lung': [0.68, 0.72, ...]
    # }
    # by defult, print the dice (1st metric) of each dataset 
    info = 'Metrics of Each Dataset:\n'
    avg_df = {}
    for dataset in datasets_labels_metrics.keys():
        avg_df[dataset] = {k:[] for k in metric_dict.keys()}    # 'TotalSegmentator(CT)': {'dice':[0.8, ...] 'nsd':[0.5, ...], ...}
        for label in datasets_labels_metrics[dataset].keys():
            avg_df[f'{dataset}, {label}'] = []
            for metric in datasets_labels_metrics[dataset][label].keys():
                label_metric = np.average(datasets_labels_metrics[dataset][label][metric])
                avg_df[f'{dataset}, {label}'].append(label_metric)  # 'TotalSegmentator, Lung': [0.68, 0.72, ...] list of num_metrics
                avg_df[dataset][metric].append(label_metric)
        avg_df[dataset] = {k:np.average(v) for k,v in avg_df[dataset].items()} # 'TotalSegmentator': {'dice':[0.8, ...] 'nsd':[0.5, ...], ...} --> 'TotalSegmentator': {'dice':0.x, 'nsd':0.x, ...}
        info += f'{dataset}  |  ' 
        for k ,v in avg_df[dataset].items():
            info += f'{v}({k})  |  '
        info += '\n'
        avg_df[dataset] = list(avg_df[dataset].values())
    avg_df = pd.DataFrame(avg_df).T
    avg_df.columns = list(metric_dict.keys())   # ['dice', 'nsd']
    avg_df.to_csv(csv_path)        
    print(info)
    
    # detailed log （nsd和dice，列为class label
    # multi-sheet, two for each dataset
    df_list = [['summary', avg_df]]
    for dataset, label_set in datasets_labels_sets.items():
        metric_df ={}
        metric_df['dice'] = {}
        metric_df['nsd'] = {}

        # create dfs like:
        # {
        #   '0.npy': [0.xx, 0.xx, ...]
        #   ......
        # }
        
        # {'COVID19':{'0.npy':{'covid19_infection':{'dice':0.8, ...} ...}, ...}
        for image_id, label_dict in samples_labels_metrics[dataset].items():
            for metric in metric_df:
                tmp = []    # one dice for each label in this dataset
                for label in label_set:
                    score = label_dict[label][metric] if label in label_dict else -1
                    tmp.append(score)
                metric_df[metric][image_id] = tmp   
        
        for metric, metric_df in metric_df.items():
            metric_df = pd.DataFrame(metric_df).T
            metric_df.columns = list(label_set)
            df_list.append([dataset+f'({metric})', metric_df])
        
    xlsx_path = csv_path.replace('.csv', '.xlsx')
    with pd.ExcelWriter(xlsx_path) as writer:
        for name, df in df_list:
            # 将每个 DataFrame 写入一个 sheet(sheet name must be < 31)
            if len(name) > 31:
                name = name[len(name)-31:]
            df.to_excel(writer, sheet_name=name, index=True)
    
    _, _ = merge('/mnt/hwfile/medai/zhaoziheng/SAM/processed_files_v4/(before NC) Statistic Files/mod_lab(72).json', xlsx_path, xlsx_path)
